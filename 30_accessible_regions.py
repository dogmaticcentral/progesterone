#!/usr/bin/python3

from Bio.Alphabet.IUPAC import unambiguous_dna
from Bio.Seq import Seq
import os
from utils.utils import *
from openpyxl import load_workbook
from math import log
import matplotlib.pyplot as plt

import numpy as np

#########################################
def read_xlsx(infile, outfile):
	outf = open(outfile, "w")
	workbook = load_workbook(filename=infile, read_only=True)
	worksheet = workbook['Table S1']

	for row in worksheet[4:worksheet.max_row]: # no zeor here, indices start from 1; first 3 rows are header
		# by inspection I know that positions 0,1,2 in rows>=5 are chrom, from, to
		# and and p_val is in pos 10
		[chrom, start, end] = [cell.value for cell in row[:3]]
		logfold_change= row[7].value
		pval = row[10].value
		if chrom != 'chr4': continue
		#if pval>1.e-2: continue
		outf.write( "\t".join([chrom, str(start), str(end), "%5.1e"%logfold_change, "%5.1e"%pval]) +"\n")
	outf.close()


def rescaled (x, offset, scale):
	return (x-offset)/scale

def read_points(infile, tad_start, tad_end, tad_length):

	inf  = open(infile,"r")
	points  = []
	values  = []
	weights = []
	for line in inf:
		[start, end, logfold_change, pval] = line.rstrip().split()[1:]
		start = int(start)
		end = int(end)
		if end<tad_start: continue
		if start>tad_end: continue
		midpoint = (start+end)/2
		point_rescaled = rescaled(midpoint, tad_start, tad_length)
		if pval==0: continue
		weight = -log(float(pval))
		points.append(point_rescaled)
		values.append(float(logfold_change))
		weights.append(weight)

	sorted_indices = [i for i in range(len(points))]
	sorted_indices.sort(key=lambda i: points[i])
	inf.close()

	np_points  = np.array(points)[sorted_indices]
	np_values    = np.array(values)[sorted_indices]
	np_weights = np.array(weights)[sorted_indices]
	#for i in range(np_points.size):
	#	print("%.3f   %5.1f   %1.f" % (np_points[i], np_values[i], np_weights[i]))

	return np_points, np_values, np_weights

#########################################
def read_tfbs_ranges(infile):
	chipseq_regions =[]
	for line in open(infile, "r"):
		if line[0]=='%': continue
		[name, chipseq_region, hic_region, hic_int_score] = line.rstrip().split("\t")[:4]
		chipseq_regions.append(chipseq_region)
	return chipseq_regions

#########################################
def main():

	gene_name = "Hand2"

	atac_file = '/storage/databases/geo/GSE104720/GSE104720_EnSC_decidualization_ATAC-seq.xlsx'
	tadfile   = "/storage/databases/encode/ENCSR551IPY/ENCFF633ORE.bed"
	ucsc_gene_regions_dir = "/storage/databases/ucsc/gene_ranges/human/hg19"
	for prerequisite in [atac_file, tadfile, ucsc_gene_regions_dir]:
		if os.path.exists(prerequisite): continue
		print(prerequisite, "not found")
		exit()

	outfile = "raw_data/chromatin_opening.tsv"

	if not os.path.exists(outfile):
		read_xlsx(atac_file, outfile)

	[chromosome, strand, gene_range] = ucsc_gene_coords(gene_name, ucsc_gene_regions_dir)
	[tad_start, tad_end] = get_tad(tadfile, chromosome, gene_range)
	tad_length = tad_end - tad_start

	# np_ are numpy arrays
	np_points, np_values, np_weights = read_points(outfile, tad_start, tad_end, tad_length)

	colors = []
	maxw = np.amax(np_weights)
	for w in np_weights:
		colors.append((0.29,0,0.5, w/maxw))

	chipseq_regions_ESR1 = read_tfbs_ranges('raw_data/hic_interactions/Hand2_ESR1_hg19.tsv')
	chipseq_regions_PGR  = read_tfbs_ranges('raw_data/hic_interactions/Hand2_PGR_hg19.tsv')

	fig, ax = plt.subplots()
	ax.axhline(0, color='black', lw=1)
	#ax.plot(np_points, np_values, 'o')
	plt.bar(np_points, np_values, width=0.01, color=colors)

	hand2_start = rescaled(gene_range[0], tad_start, tad_length)
	hand2_end  = rescaled(gene_range[1], tad_start, tad_length)
	ax.hlines(y=0.0, xmin=hand2_start-0.002, xmax=hand2_end+0.002, linewidth=30, color='b')

	for chipseq_reg in chipseq_regions_ESR1:
		[start, end] = [rescaled(int(p), tad_start, tad_length) for p in chipseq_reg.split("_")]
		ax.hlines(y=0.0, xmin=start-0.002, xmax=end+0.002, linewidth=30, color='r')

	for chipseq_reg in chipseq_regions_PGR:
		[start, end] = [rescaled(int(p), tad_start, tad_length) for p in chipseq_reg.split("_")]
		ax.hlines(y=0.0, xmin=start-0.002, xmax=end+0.002, linewidth=30, color='g')

	plt.show()

	return

#########################################
if __name__== '__main__':
	main()
