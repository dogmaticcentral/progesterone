#!/usr/bin/python3

#
# This file is part of Progesterone pipeline.
#
# Progesterone pipeline  is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# Progesterone pipeline is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Progesterone pipeline.  If not, see <https://www.gnu.org/licenses/>.
#

from utils.mysqldb import *
from utils.utils import *
from openpyxl import load_workbook


#########################################
def read_xlsx(infile, tgt_chromosome, tad_start, tad_end):

	region_changes = []

	workbook = load_workbook(filename=infile, read_only=True)
	worksheet = workbook['Table S1']

	for row in worksheet[4:worksheet.max_row]: # no zero here, indices start from 1; first 3 rows are header
		# by inspection I know that positions 0,1,2 in rows>=5 are chrom, from, to
		# and and p_val is in pos 10
		[chrom, start, end] = [cell.value for cell in row[:3]]
		if chrom != tgt_chromosome: continue
		if end<tad_start or start>tad_end: continue
		logfold_change= row[7].value
		pval = row[10].value
		#if pval>1.e-2: continue
		region_changes.append([start, end,  logfold_change,  pval])

	return region_changes

def rescaled (x, offset, scale):
	return (x-offset)/scale

#########################################
def main():

	species   = "human"
	assembly  = "hg19"
	gene_name = "Hand2"
	tad_external_exp_id = "ENCFF633ORE"
	atac_file = '/storage/databases/geo/GSE104720/GSE104720_EnSC_decidualization_ATAC-seq.xlsx'
	pubmed_id = '29259032'
	conf_file = "/home/ivana/.mysql_conf"

	for prerequisite in [atac_file, conf_file]:
		if os.path.exists(prerequisite): continue
		print(prerequisite, "not found")
		exit()

	db = connect_to_mysql(conf_file)
	cursor = db.cursor()
	switch_to_db(cursor,'progesterone')
	search_db(cursor,"set autocommit=1")

	# find xref_id for the experimental data file
	atac_xref_id = store_xref (cursor, 'pubmed', pubmed_id)
	tad_xref_id  = get_xref_id(db,cursor,tad_external_exp_id)
	# find gene coordinates
	[chromosome, strand, gene_start, gene_end] = get_gene_coords(db,cursor,gene_name,assembly)
	# tad?
	[tad_start, tad_end] = get_tad_region(db, cursor, tad_xref_id, chromosome, gene_start, gene_end)

	#read atac regions
	region_changes = read_xlsx(atac_file, chromosome, tad_start, tad_end)

	# store atac regions
	for rc in region_changes:
		[start, end, logfold_change, pval] = rc
		fields  = {'species':species, 'chromosome':chromosome, 'assembly':assembly, 'rtype':'atacseq',
					'rfrom':start, 'rto':end, 'xref_id':atac_xref_id}
		atac_region_id = store_or_update(cursor, 'regions', fields, None)
		# store info about the atac region
		fixed_fields  = {'atac_region_id':atac_region_id}
		update_fields = {'logfold_change':logfold_change, 'pval':pval}
		atac_id = store_or_update (cursor, 'atac_acc_changes', fixed_fields, update_fields)

	cursor.close()
	db.close()


	return

#########################################
if __name__== '__main__':
	main()
